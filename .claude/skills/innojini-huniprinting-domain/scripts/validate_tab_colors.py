#!/usr/bin/env python3
"""
Tab Color Consistency Validation Script
탭 색상 일관성 검증 스크립트

Usage:
    python validate_tab_colors.py workbook.xlsx

Validates that sheet tab colors match expected conventions.
"""

import sys
import json
from openpyxl import load_workbook

# 상품마스터 예상 탭 색상
MASTER_EXPECTED_COLORS = {
    "MAP": "FFE36C09",           # Orange - Master/Mapping
    "!디지털인쇄용지": "FFFF0000", # Red - Reference Data
    "비즈하우스": "FF6AA84F",     # Green - External Channel
    "후지필름": "FF6AA84F",       # Green - External Channel
    "스토어": "FF00B050",         # Green (alt) - External Channel
}

# 인쇄가격표 예상 탭 색상
PRICE_EXPECTED_COLORS = {
    "디지털용지": "FFFF9900",      # Orange - Core Price Table
    "디지털출력비": "FFFF9900",    # Orange - Core Price Table
    "후가공": "FFFF9900",          # Orange - Core Price Table
    "옵션결합상품": "FFFF0000",    # Red - Reference Data
    "디지털출력비가수정": "FF00FF00",  # Green - Work in Progress
}

# 색상 의미 매핑
COLOR_MEANINGS = {
    "FFE36C09": "🟠 Master/Mapping",
    "FFFF0000": "🔴 Reference Data",
    "FF6AA84F": "🟢 External Channel",
    "FF00B050": "🟢 External Channel (alt)",
    "FFFF9900": "🟠 Core Price Table",
    "FF00FF00": "🟢 Work in Progress",
    None: "⬜ Default (Core Product/General)",
}


def get_color_meaning(color_hex):
    """색상 코드의 의미 반환"""
    return COLOR_MEANINGS.get(color_hex, f"❓ Unknown ({color_hex})")


def validate_tab_colors(filepath, expected_colors=None):
    """탭 색상 일관성 검증"""
    
    wb = load_workbook(filepath)
    
    results = {
        "filepath": filepath,
        "total_sheets": len(wb.sheetnames),
        "sheets": [],
        "mismatches": [],
        "status": "success",
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tab_color = ws.sheet_properties.tabColor
        actual_hex = tab_color.rgb if tab_color else None
        
        sheet_info = {
            "name": sheet_name,
            "color_hex": actual_hex,
            "color_meaning": get_color_meaning(actual_hex),
        }
        
        # 예상 색상과 비교
        if expected_colors and sheet_name in expected_colors:
            expected_hex = expected_colors[sheet_name]
            sheet_info["expected_hex"] = expected_hex
            sheet_info["expected_meaning"] = get_color_meaning(expected_hex)
            
            if actual_hex != expected_hex:
                sheet_info["match"] = False
                results["mismatches"].append({
                    "sheet": sheet_name,
                    "expected": expected_hex,
                    "actual": actual_hex,
                })
            else:
                sheet_info["match"] = True
        
        results["sheets"].append(sheet_info)
    
    wb.close()
    
    # 상태 결정
    if results["mismatches"]:
        results["status"] = "warning"
        results["message"] = f"탭 색상 불일치 {len(results['mismatches'])}개"
    else:
        results["message"] = "탭 색상 일관성 정상"
    
    return results


def detect_file_type(filepath):
    """파일 유형 자동 감지"""
    wb = load_workbook(filepath, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    if "MAP" in sheet_names or "!디지털인쇄용지" in sheet_names:
        return "master", MASTER_EXPECTED_COLORS
    elif "디지털용지" in sheet_names or "디지털출력비" in sheet_names:
        return "price", PRICE_EXPECTED_COLORS
    else:
        return "unknown", {}


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_tab_colors.py workbook.xlsx")
        sys.exit(1)
    
    filepath = sys.argv[1]
    
    # 파일 유형 감지
    file_type, expected_colors = detect_file_type(filepath)
    
    results = validate_tab_colors(filepath, expected_colors)
    results["file_type"] = file_type
    
    print(json.dumps(results, ensure_ascii=False, indent=2))
    
    # 종료 코드
    if results["status"] == "warning":
        sys.exit(1)
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()
