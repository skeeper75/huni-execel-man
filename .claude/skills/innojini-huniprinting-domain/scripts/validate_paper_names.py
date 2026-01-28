#!/usr/bin/env python3
"""
Paper Name Cross-Validation Script
용지명 교차 검증 스크립트

Usage:
    python validate_paper_names.py master.xlsx pricing.xlsx

Validates paper name consistency between product master and price table.
"""

import sys
import json
from openpyxl import load_workbook

# 약어 매핑 테이블
PAPER_ABBREVIATIONS = {
    "WH": "울트라화이트",
    "T3절": "3절",
    "T4절": "4절",
    "T국전": "국전",
}

def normalize_paper_name(name):
    """용지명 정규화"""
    if not name:
        return ""
    
    name = str(name).strip()
    
    # 1. 특수 마커 제거
    for marker in ["★", "●", "▶︎", "▶"]:
        name = name.replace(marker, "")
    
    # 2. 약어 치환
    for abbr, full in PAPER_ABBREVIATIONS.items():
        name = name.replace(abbr, full)
    
    # 3. 공백 정규화
    name = " ".join(name.split())
    
    return name.strip()


def extract_master_papers(filepath):
    """상품마스터에서 용지명 추출"""
    wb = load_workbook(filepath, data_only=True)
    
    papers = {}
    
    # !디지털인쇄용지 시트에서 추출
    if '!디지털인쇄용지' in wb.sheetnames:
        ws = wb['!디지털인쇄용지']
        for row in range(2, ws.max_row + 1):
            paper_name = ws.cell(row=row, column=2).value  # 종이명 컬럼
            if paper_name and str(paper_name).strip():
                original = str(paper_name).strip()
                normalized = normalize_paper_name(original)
                papers[normalized] = original
    
    wb.close()
    return papers


def extract_price_papers(filepath):
    """인쇄가격표에서 용지명 추출"""
    wb = load_workbook(filepath, data_only=True)
    
    papers = {}
    
    # 디지털용지 시트에서 추출
    if '디지털용지' in wb.sheetnames:
        ws = wb['디지털용지']
        for row in range(3, ws.max_row + 1):
            paper_name = ws.cell(row=row, column=2).value  # 종이명 컬럼
            if paper_name and str(paper_name).strip():
                original = str(paper_name).strip()
                normalized = normalize_paper_name(original)
                papers[normalized] = original
    
    wb.close()
    return papers


def validate_paper_names(master_path, price_path):
    """용지명 교차 검증"""
    
    master_papers = extract_master_papers(master_path)
    price_papers = extract_price_papers(price_path)
    
    master_set = set(master_papers.keys())
    price_set = set(price_papers.keys())
    
    common = master_set & price_set
    master_only = master_set - price_set
    price_only = price_set - master_set
    
    results = {
        "status": "success" if not master_only else "warning",
        "total_master": len(master_papers),
        "total_price": len(price_papers),
        "common_count": len(common),
        "master_only_count": len(master_only),
        "price_only_count": len(price_only),
        "master_only": [
            {"normalized": n, "original": master_papers[n]} 
            for n in sorted(master_only)
        ],
        "price_only": [
            {"normalized": n, "original": price_papers[n]} 
            for n in sorted(price_only)
        ],
    }
    
    # 심각도 판정
    if master_only:
        results["severity"] = "P1"
        results["message"] = f"마스터에만 있는 용지 {len(master_only)}개 - 가격 조회 불가 위험"
    elif price_only:
        results["severity"] = "P2"
        results["message"] = f"가격표에만 있는 용지 {len(price_only)}개 - 사용되지 않는 데이터"
    else:
        results["severity"] = "OK"
        results["message"] = "용지명 정합성 정상"
    
    return results


def main():
    if len(sys.argv) < 3:
        print("Usage: python validate_paper_names.py master.xlsx pricing.xlsx")
        sys.exit(1)
    
    master_path = sys.argv[1]
    price_path = sys.argv[2]
    
    results = validate_paper_names(master_path, price_path)
    
    print(json.dumps(results, ensure_ascii=False, indent=2))
    
    # 종료 코드
    if results["severity"] == "P1":
        sys.exit(1)
    elif results["severity"] == "P2":
        sys.exit(0)  # 경고만, 실패 아님
    else:
        sys.exit(0)


if __name__ == "__main__":
    main()
