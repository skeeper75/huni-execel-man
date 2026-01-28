#!/usr/bin/env python3
"""
Price spreadsheet validation script.

Usage:
    python validate_price_sheet.py <excel_file>

Returns JSON with validation results.
Exit code 0 = PASS, 1 = FAIL
"""

import sys
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
    import pandas as pd
except ImportError as e:
    print(json.dumps({
        "status": "ERROR",
        "message": f"Missing dependency: {e}. Install with: pip install openpyxl pandas"
    }))
    sys.exit(2)


ERROR_VALUES = ['#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NULL!', '#NUM!']


def detect_formula_errors(filepath):
    """Detect Excel formula errors."""
    errors = []
    
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return [{"type": "FILE_ERROR", "message": str(e), "severity": "CRITICAL"}]
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in ERROR_VALUES:
                    errors.append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "error": str(cell.value),
                        "severity": "CRITICAL"
                    })
    
    return errors


def detect_negative_prices(filepath, threshold=100):
    """Detect negative values that appear to be prices."""
    warnings = []
    
    try:
        xl = pd.ExcelFile(filepath)
    except Exception as e:
        return [{"type": "FILE_ERROR", "message": str(e)}]
    
    for sheet_name in xl.sheet_names:
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        except Exception:
            continue
            
        for col_idx, col in enumerate(df.columns):
            for row_idx, val in df[col].items():
                try:
                    num_val = float(val)
                    if num_val < 0 and abs(num_val) >= threshold:
                        warnings.append({
                            "sheet": sheet_name,
                            "row": int(row_idx) + 1,
                            "col": int(col_idx) + 1,
                            "value": num_val,
                            "issue": "Potential negative price",
                            "severity": "HIGH"
                        })
                except (ValueError, TypeError):
                    pass
    
    return warnings


def analyze_statistics(filepath):
    """Generate basic statistics."""
    stats = {}
    
    try:
        xl = pd.ExcelFile(filepath)
    except Exception:
        return stats
    
    for sheet_name in xl.sheet_names:
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            numeric_cols = df.select_dtypes(include=['number']).columns
            
            sheet_stats = {
                "rows": len(df),
                "columns": len(df.columns),
                "numeric_columns": len(numeric_cols),
                "null_count": int(df.isnull().sum().sum())
            }
            
            if len(numeric_cols) > 0:
                all_numeric = df[numeric_cols].values.flatten()
                all_numeric = all_numeric[~pd.isna(all_numeric)]
                if len(all_numeric) > 0:
                    sheet_stats["min_value"] = float(min(all_numeric))
                    sheet_stats["max_value"] = float(max(all_numeric))
            
            stats[sheet_name] = sheet_stats
        except Exception:
            stats[sheet_name] = {"error": "Could not analyze"}
    
    return stats


def run_validation(filepath):
    """Run complete validation."""
    results = {
        "file": str(filepath),
        "status": "PASS",
        "errors": [],
        "warnings": [],
        "statistics": {},
        "summary": {}
    }
    
    # Formula errors
    formula_errors = detect_formula_errors(filepath)
    if formula_errors:
        results["status"] = "FAIL"
        results["errors"].extend(formula_errors)
    
    # Negative prices
    negative_prices = detect_negative_prices(filepath)
    results["warnings"].extend(negative_prices)
    
    # Statistics
    results["statistics"] = analyze_statistics(filepath)
    
    # Summary
    results["summary"] = {
        "formula_errors": len([e for e in results["errors"] if e.get("error") in ERROR_VALUES]),
        "negative_price_warnings": len(negative_prices),
        "sheets_analyzed": len(results["statistics"]),
        "total_issues": len(results["errors"]) + len(results["warnings"])
    }
    
    return results


def main():
    if len(sys.argv) < 2:
        print(json.dumps({
            "status": "ERROR",
            "message": "Usage: python validate_price_sheet.py <excel_file>"
        }))
        sys.exit(2)
    
    filepath = Path(sys.argv[1])
    
    if not filepath.exists():
        print(json.dumps({"status": "ERROR", "message": f"File not found: {filepath}"}))
        sys.exit(2)
    
    if not filepath.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
        print(json.dumps({"status": "ERROR", "message": f"Not an Excel file: {filepath}"}))
        sys.exit(2)
    
    results = run_validation(str(filepath))
    print(json.dumps(results, indent=2, ensure_ascii=False))
    
    sys.exit(0 if results["status"] == "PASS" else 1)


if __name__ == '__main__':
    main()
