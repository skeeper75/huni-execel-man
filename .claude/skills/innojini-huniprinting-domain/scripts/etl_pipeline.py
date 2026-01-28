#!/usr/bin/env python3
"""
Printing product master ETL pipeline.

Usage:
    python etl_pipeline.py <excel_file> [options]

Options:
    --output DIR      Output directory (default: ./output)
    --format FORMAT   Output format: json, sqlite, both (default: json)
    --sheets SHEETS   Sheets to process (comma-separated)
    --validate-only   Run validation only
"""

import sys
import os
import json
import re
import argparse
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError as e:
    print(json.dumps({
        "status": "ERROR",
        "message": f"Missing dependency: {e}. Install with: pip install pandas openpyxl"
    }))
    sys.exit(2)


MARKERS = {
    '▶︎': 'category_header',
    '▶': 'category_header',
    '★': 'new_product',
    '●': 'applicable',
    '#': 'reference_tag',
}

PRODUCT_SHEETS = [
    '디지털인쇄', '스티커', '책자', '포토북', '캘린더',
    '실사', '아크릴', '굿즈', '문구(노트)', '상품악세사리'
]


def extract_workbook(filepath):
    """Extract workbook with metadata."""
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return {"error": f"Failed to load: {e}"}
    
    extracted = {
        'sheets': {},
        'metadata': {
            'filename': str(filepath),
            'sheet_count': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'tab_colors': {}
        }
    }
    
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            if ws.sheet_properties.tabColor:
                extracted['metadata']['tab_colors'][sheet_name] = ws.sheet_properties.tabColor.rgb
            
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            header_row = detect_header_row(df)
            
            extracted['sheets'][sheet_name] = {
                'data': df,
                'shape': df.shape,
                'header_row': header_row
            }
        except Exception as e:
            extracted['sheets'][sheet_name] = {'error': str(e)}
    
    return extracted


def detect_header_row(df, max_rows=10):
    """Detect header row."""
    for i in range(min(max_rows, len(df))):
        row = df.iloc[i]
        str_count = sum(1 for v in row if isinstance(v, str) and len(str(v)) > 0)
        if str_count >= 3:
            return i
    return 0


def parse_mes_code(code):
    """Parse MES code."""
    if not code or pd.isna(code):
        return None
    pattern = r'^(\d{3})-(\d{4})$'
    match = re.match(pattern, str(code).strip())
    if match:
        return {'full_code': str(code).strip(), 'category': match.group(1), 'sequence': match.group(2)}
    return None


def parse_size(size_str):
    """Parse size string."""
    if not size_str or pd.isna(size_str):
        return None
    size_str = str(size_str).strip()
    for sep in ['x', 'X', '×', '*']:
        if sep in size_str:
            parts = size_str.split(sep)
            if len(parts) == 2:
                try:
                    return {'width': float(parts[0].strip()), 'height': float(parts[1].strip())}
                except ValueError:
                    pass
    return {'raw': size_str}


def parse_markers(value):
    """Parse markers from value."""
    if not isinstance(value, str):
        return {'value': value, 'markers': []}
    markers = []
    clean_value = value
    for marker, meaning in MARKERS.items():
        if marker in value:
            markers.append(meaning)
            clean_value = clean_value.replace(marker, '').strip()
    return {'value': clean_value, 'markers': markers, 'original': value}


def transform_to_entities(extracted_data, target_sheets=None):
    """Transform to entities."""
    if 'error' in extracted_data:
        return {'error': extracted_data['error']}
    
    entities = {'papers': [], 'categories': [], 'products': [], 'summary': {}}
    
    # Paper master
    for name in extracted_data['sheets'].keys():
        if '용지' in name or 'paper' in name.lower():
            if 'data' in extracted_data['sheets'].get(name, {}):
                entities['papers'] = extract_paper_master(extracted_data['sheets'][name])
            break
    
    # Categories from MAP
    if 'MAP' in extracted_data['sheets'] and 'data' in extracted_data['sheets']['MAP']:
        entities['categories'] = extract_categories(extracted_data['sheets']['MAP'])
    
    # Products
    sheets_to_process = target_sheets if target_sheets else PRODUCT_SHEETS
    for sheet_name in sheets_to_process:
        if sheet_name in extracted_data['sheets']:
            sheet_data = extracted_data['sheets'][sheet_name]
            if 'data' in sheet_data:
                entities['products'].extend(extract_products(sheet_data, sheet_name))
    
    entities = create_references(entities)
    entities['summary'] = {
        'paper_count': len(entities['papers']),
        'category_count': len(entities['categories']),
        'product_count': len(entities['products']),
        'processed_sheets': [s for s in sheets_to_process if s in extracted_data['sheets']]
    }
    
    return entities


def extract_paper_master(sheet_data):
    """Extract paper master."""
    df = sheet_data['data']
    header_row = sheet_data['header_row']
    headers = df.iloc[header_row].tolist()
    papers = []
    
    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx]
        if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
            continue
        
        paper = {'paper_id': f"PAPER_{idx:04d}", 'name': str(row.iloc[0]).strip(), 'applicable_products': []}
        
        for col_idx, header in enumerate(headers):
            if pd.isna(header):
                continue
            header_str = str(header).strip()
            value = row.iloc[col_idx] if col_idx < len(row) else None
            
            if '평량' in header_str:
                paper['gram'] = value
            elif '전지' in header_str or '규격' in header_str:
                paper['full_sheet_size'] = value
            elif '가' in header_str and '격' in header_str:
                paper['price'] = value
            elif header_str.startswith('#') and value == '●':
                paper['applicable_products'].append(header_str[1:])
        
        papers.append(paper)
    
    return papers


def extract_categories(sheet_data):
    """Extract categories from MAP."""
    df = sheet_data['data']
    categories = []
    
    for col_idx in range(min(len(df.columns), 20)):
        col = df.iloc[:, col_idx]
        parent_id = None
        
        for row_idx, value in enumerate(col):
            if pd.isna(value):
                continue
            parsed = parse_markers(str(value))
            
            if 'category_header' in parsed['markers']:
                cat_id = f"CAT_{col_idx:02d}"
                categories.append({
                    'id': cat_id, 'code': f"{col_idx+1:02d}", 'name': parsed['value'],
                    'parent_id': None, 'is_header': True
                })
                parent_id = cat_id
            elif parsed['value'] and parent_id:
                sub_id = f"CAT_{col_idx:02d}_{row_idx:03d}"
                categories.append({
                    'id': sub_id, 'code': f"{col_idx+1:02d}-{row_idx:03d}", 'name': parsed['value'],
                    'parent_id': parent_id, 'is_new': 'new_product' in parsed['markers']
                })
    
    return categories


def extract_products(sheet_data, sheet_name):
    """Extract products."""
    df = sheet_data['data']
    header_row = sheet_data['header_row']
    headers = df.iloc[header_row].tolist()
    
    col_map = {}
    for idx, h in enumerate(headers):
        if pd.isna(h):
            continue
        h_str = str(h).strip().lower()
        if 'mes' in h_str or 'item_cd' in h_str or h_str == 'code':
            col_map['mes_code'] = idx
        elif '상품명' in h_str or 'name' in h_str:
            col_map['name'] = idx
        elif '사이즈옵션' in h_str:
            col_map['size_option'] = idx
        elif '재단' in h_str and '사이즈' in h_str:
            col_map['trim_size'] = idx
        elif '작업' in h_str and '사이즈' in h_str:
            col_map['work_size'] = idx
        elif '파일크기' in h_str:
            col_map['work_size'] = idx
        elif '블리드' in h_str:
            col_map['bleed'] = idx
        elif '판수' in h_str:
            col_map['pansu'] = idx
        elif '종이' in h_str or 'material' in h_str:
            col_map['paper'] = idx
    
    products = []
    for idx in range(header_row + 1, len(df)):
        row = df.iloc[idx]
        mes_code = row.iloc[col_map['mes_code']] if 'mes_code' in col_map else None
        
        if pd.isna(mes_code) or str(mes_code).strip() == '':
            continue
        
        parsed_code = parse_mes_code(mes_code)
        product = {
            'mes_code': str(mes_code).strip(),
            'category_code': parsed_code['category'] if parsed_code else None,
            'source_sheet': sheet_name,
            'row_index': idx
        }
        
        for field, col_idx in col_map.items():
            if field == 'mes_code':
                continue
            value = row.iloc[col_idx] if col_idx < len(row) else None
            
            if field in ['trim_size', 'work_size']:
                product[field] = parse_size(value)
            elif field == 'name':
                parsed = parse_markers(str(value) if not pd.isna(value) else '')
                product['name'] = parsed['value']
                product['is_new'] = 'new_product' in parsed['markers']
            else:
                product[field] = value if not pd.isna(value) else None
        
        products.append(product)
    
    return products


def create_references(entities):
    """Create references."""
    paper_map = {p['name']: p['paper_id'] for p in entities['papers'] if p.get('name')}
    for product in entities['products']:
        paper_name = product.get('paper')
        product['paper_id'] = paper_map.get(paper_name) if paper_name else None
    return entities


def export_to_json(entities, output_dir):
    """Export to JSON."""
    os.makedirs(output_dir, exist_ok=True)
    outputs = {}
    
    for entity_name in ['papers', 'categories', 'products']:
        data = entities.get(entity_name, [])
        filename = f"{output_dir}/{entity_name}.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        outputs[entity_name] = {'filename': filename, 'record_count': len(data)}
    
    meta = {'exported_at': datetime.now().isoformat(), 'summary': entities.get('summary', {}), 'outputs': outputs}
    with open(f"{output_dir}/export_meta.json", 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    
    return outputs


def export_to_sqlite(entities, db_path):
    """Export to SQLite."""
    import sqlite3
    conn = sqlite3.connect(db_path)
    
    conn.execute('CREATE TABLE IF NOT EXISTS papers (paper_id TEXT PRIMARY KEY, name TEXT, gram TEXT, full_sheet_size TEXT, price REAL)')
    conn.execute('CREATE TABLE IF NOT EXISTS products (mes_code TEXT PRIMARY KEY, name TEXT, category_code TEXT, paper_id TEXT, source_sheet TEXT, size_option TEXT, trim_size_w REAL, trim_size_h REAL, work_size_w REAL, work_size_h REAL, bleed REAL, pansu INTEGER, is_new INTEGER)')
    
    for paper in entities['papers']:
        conn.execute('INSERT OR REPLACE INTO papers VALUES (?, ?, ?, ?, ?)',
            (paper.get('paper_id'), paper.get('name'), str(paper.get('gram')) if paper.get('gram') else None, paper.get('full_sheet_size'), paper.get('price')))
    
    for product in entities['products']:
        trim = product.get('trim_size') or {}
        work = product.get('work_size') or {}
        conn.execute('INSERT OR REPLACE INTO products VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (product.get('mes_code'), product.get('name'), product.get('category_code'), product.get('paper_id'), product.get('source_sheet'), product.get('size_option'), trim.get('width'), trim.get('height'), work.get('width'), work.get('height'), product.get('bleed'), product.get('pansu'), 1 if product.get('is_new') else 0))
    
    conn.commit()
    conn.close()
    return db_path


def validate_entities(entities):
    """Validate entities."""
    errors, warnings = [], []
    
    for product in entities['products']:
        if not parse_mes_code(product['mes_code']):
            errors.append({'type': 'INVALID_MES_CODE', 'product': product.get('name'), 'value': product['mes_code']})
    
    mes_codes = [p['mes_code'] for p in entities['products']]
    for dup in [c for c in set(mes_codes) if mes_codes.count(c) > 1]:
        warnings.append({'type': 'DUPLICATE_MES_CODE', 'value': dup, 'count': mes_codes.count(dup)})
    
    paper_ids = {p['paper_id'] for p in entities['papers']}
    for product in entities['products']:
        pid = product.get('paper_id')
        if pid and pid not in paper_ids:
            warnings.append({'type': 'ORPHAN_PAPER_REF', 'product': product.get('name'), 'paper_id': pid})
    
    return {'errors': errors, 'warnings': warnings, 'valid': len(errors) == 0}


def main():
    parser = argparse.ArgumentParser(description='Printing product master ETL')
    parser.add_argument('excel_file', help='Input Excel file')
    parser.add_argument('--output', '-o', default='./output', help='Output directory')
    parser.add_argument('--format', '-f', choices=['json', 'sqlite', 'both'], default='json')
    parser.add_argument('--sheets', '-s', help='Sheets to process (comma-separated)')
    parser.add_argument('--validate-only', action='store_true', help='Run validation only')
    
    args = parser.parse_args()
    filepath = Path(args.excel_file)
    
    if not filepath.exists():
        print(json.dumps({"status": "ERROR", "message": f"File not found: {filepath}"}))
        sys.exit(2)
    
    print(f"📂 Loading: {filepath}", file=sys.stderr)
    extracted = extract_workbook(filepath)
    
    if 'error' in extracted:
        print(json.dumps({"status": "ERROR", "message": extracted['error']}))
        sys.exit(2)
    
    print(f"   Found {len(extracted['sheets'])} sheets", file=sys.stderr)
    
    target_sheets = args.sheets.split(',') if args.sheets else None
    entities = transform_to_entities(extracted, target_sheets)
    
    print(f"   Papers: {entities['summary']['paper_count']}", file=sys.stderr)
    print(f"   Products: {entities['summary']['product_count']}", file=sys.stderr)
    
    validation = validate_entities(entities)
    
    if validation['errors']:
        print(f"\n❌ Errors: {len(validation['errors'])}", file=sys.stderr)
        for err in validation['errors'][:5]:
            print(f"   - {err['type']}: {err.get('value')}", file=sys.stderr)
    
    if validation['warnings']:
        print(f"\n⚠️  Warnings: {len(validation['warnings'])}", file=sys.stderr)
        for warn in validation['warnings'][:5]:
            print(f"   - {warn['type']}: {warn.get('value')}", file=sys.stderr)
    
    if args.validate_only:
        print(json.dumps(validation, indent=2))
        sys.exit(0 if validation['valid'] else 1)
    
    print(f"\n💾 Exporting...", file=sys.stderr)
    os.makedirs(args.output, exist_ok=True)
    
    if args.format in ['json', 'both']:
        export_to_json(entities, args.output)
        print(f"   JSON: {args.output}/", file=sys.stderr)
    
    if args.format in ['sqlite', 'both']:
        db_path = export_to_sqlite(entities, f"{args.output}/products.db")
        print(f"   SQLite: {db_path}", file=sys.stderr)
    
    print("\n✅ Done!", file=sys.stderr)
    print(json.dumps({"status": "SUCCESS", "summary": entities['summary'], "validation": validation}, indent=2))


if __name__ == '__main__':
    main()
